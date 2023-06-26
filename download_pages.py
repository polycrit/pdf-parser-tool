import pdfkit
from pathlib import Path
import requests
import openpyxl
import os

wb = openpyxl.load_workbook("Regulations.xlsx")
ws = wb["Regulations"]
# This will fail if there is no hyperlink to target
arr_links = []

for i in range(1, 37):
    try:
        name = str(ws.cell(row=i, column=2).value) + "_en"
        link = ws.cell(row=i, column=8).hyperlink.target
        arr_links.append({"name": name, "link": link})
    except AttributeError:
        print("no hyperlink in row " + str(i) + ", column 8")

    try:
        name = str(ws.cell(row=i, column=2).value) + "_de"
        link = ws.cell(row=i, column=9).hyperlink.target
        arr_links.append({"name": name, "link": link})
    except AttributeError:
        print("no hyperlink in row " + str(i) + ", column 9")


for doc in arr_links:
    res = os.path.join("data", doc["name"] + ".pdf")
    try:
        pdfkit.from_url(doc["link"], res)
    except OSError:
        print("Couldn't scrap the page. Trying to download it...")
        response = requests.get(doc["link"])
        Path(res).write_bytes(response.content)
